import io
from .base import BaseParser

class XLSXParser(BaseParser):
    def parse(self, file_content: bytes) -> dict:
        text = self.extract_text(file_content)
        meta = self.extract_metadata(file_content)
        tables = self.extract_tables(file_content)
        images = self.extract_images(file_content)
        structure = self.extract_structure(file_content)

        words = len(text.split())
        chars = len(text)
        
        return {
            "title": meta.get("title", ""),
            "metadata": meta,
            "sections": structure,
            "paragraphs": [{"text": p.strip()} for p in text.split("\n\n") if p.strip()],
            "tables": tables,
            "images": images,
            "links": [],
            "language": "en",
            "statistics": {
                "word_count": words,
                "character_count": chars,
                "page_count": len(meta.get("sheets", ["Sheet1"])),
                "table_count": len(tables),
                "image_count": len(images)
            }
        }

    def extract_text(self, file_content: bytes) -> str:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            lines = []
            for sheet in wb.worksheets:
                lines.append(f"Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    row_str = " | ".join([str(val) for val in row if val is not None])
                    if row_str.strip():
                        lines.append(row_str)
            return "\n\n".join(lines)
        except ImportError:
            return "openpyxl not installed. Return empty fallback."
        except Exception as e:
            return f"Error parsing XLSX: {str(e)}"

    def extract_tables(self, file_content: bytes) -> list[dict]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            tables_list = []
            for sheet in wb.worksheets:
                grid = []
                for row in sheet.iter_rows(values_only=True):
                    grid.append([str(val) if val is not None else "" for val in row])
                tables_list.append({
                    "sheet_name": sheet.title,
                    "data": grid
                })
            return tables_list
        except Exception:
            return []

    def extract_images(self, file_content: bytes) -> list[dict]:
        # openpyxl has wb.worksheets[i]._images
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            images_list = []
            for sheet in wb.worksheets:
                if hasattr(sheet, "_images"):
                    for img in sheet._images:
                        images_list.append({
                            "sheet": sheet.title,
                            "anchor": str(img.anchor) if hasattr(img, "anchor") else ""
                        })
            return images_list
        except Exception:
            return []

    def extract_metadata(self, file_content: bytes) -> dict:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            return {
                "sheets": wb.sheetnames,
                "creator": wb.properties.creator or ""
            }
        except Exception:
            return {"sheets": ["Sheet1"]}

    def extract_structure(self, file_content: bytes) -> list[dict]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_content))
            return [{"level": "1", "title": name} for name in wb.sheetnames]
        except Exception:
            return []
